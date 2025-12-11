from openpyxl import load_workbook
wb = load_workbook("templates/Ficha_KYC_1 .xlsm", keep_vba=True)
print(wb.sheetnames)
